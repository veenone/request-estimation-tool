"""Excel report generation with openpyxl.

Generates a multi-sheet workbook per SPEC §5.3:
  Sheet 1 – Summary: Key metrics, request details, and totals
  Sheet 2 – Task Breakdown: Full task list with all calculation columns
  Sheet 3 – DUT-Profile Matrix: Grid showing combinations
  Sheet 4 – Team Allocation: Tester assignments with hours per person
  Sheet 5 – PR Fixes: PR breakdown by complexity
  Sheet 6 – Reference Data: Historical project comparison
"""

from io import BytesIO
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# ── Styling constants ────────────────────────────────────

HEADER_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
TITLE_FONT = Font(name="Calibri", bold=True, size=14, color="2F5496")
SUBTITLE_FONT = Font(name="Calibri", bold=True, size=11, color="2F5496")
LABEL_FONT = Font(name="Calibri", bold=True, size=10)
VALUE_FONT = Font(name="Calibri", size=10)
FEASIBLE_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
AT_RISK_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
NOT_FEASIBLE_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def _style_header_row(ws: Any, row: int, col_count: int) -> None:
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER


def _auto_width(ws: Any) -> None:
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 3, 50)


def _feasibility_fill(status: str) -> PatternFill:
    if status == "FEASIBLE":
        return FEASIBLE_FILL
    elif status == "AT_RISK":
        return AT_RISK_FILL
    return NOT_FEASIBLE_FILL


# ── Data structure ───────────────────────────────────────

class ExcelReportData:
    """Container for all data needed to generate the Excel report."""

    def __init__(
        self,
        project_name: str,
        estimation_number: str,
        project_type: str,
        created_by: str | None = None,
        created_at: str | None = None,
        # Request info
        request_number: str | None = None,
        requester_name: str | None = None,
        business_unit: str | None = None,
        priority: str | None = None,
        # Estimation totals
        dut_count: int = 0,
        profile_count: int = 0,
        dut_profile_combinations: int = 0,
        pr_fix_count: int = 0,
        expected_delivery: str | None = None,
        total_tester_hours: float = 0,
        total_leader_hours: float = 0,
        pr_fix_hours: float = 0,
        study_hours: float = 0,
        buffer_hours: float = 0,
        grand_total_hours: float = 0,
        grand_total_days: float = 0,
        feasibility_status: str = "FEASIBLE",
        capacity_hours: float = 0,
        utilization_pct: float = 0,
        # Task breakdown
        tasks: list[dict] | None = None,
        # DUT and profile details
        dut_types: list[dict] | None = None,
        profiles: list[dict] | None = None,
        dut_profile_matrix: list[list[int]] | None = None,
        # Team
        team_members: list[dict] | None = None,
        team_size: int = 0,
        has_leader: bool = False,
        # PR fixes breakdown
        pr_simple: int = 0,
        pr_medium: int = 0,
        pr_complex: int = 0,
        # Reference projects
        reference_projects: list[dict] | None = None,
        # Risk flags
        risk_flags: list[str] | None = None,
        risk_messages: list[str] | None = None,
        # PR details
        pr_details: list[dict] | None = None,
    ):
        self.project_name = project_name
        self.estimation_number = estimation_number
        self.project_type = project_type
        self.created_by = created_by or ""
        self.created_at = created_at or ""
        self.request_number = request_number or ""
        self.requester_name = requester_name or ""
        self.business_unit = business_unit or ""
        self.priority = priority or ""
        self.dut_count = dut_count
        self.profile_count = profile_count
        self.dut_profile_combinations = dut_profile_combinations
        self.pr_fix_count = pr_fix_count
        self.expected_delivery = expected_delivery or ""
        self.total_tester_hours = total_tester_hours
        self.total_leader_hours = total_leader_hours
        self.pr_fix_hours = pr_fix_hours
        self.study_hours = study_hours
        self.buffer_hours = buffer_hours
        self.grand_total_hours = grand_total_hours
        self.grand_total_days = grand_total_days
        self.feasibility_status = feasibility_status
        self.capacity_hours = capacity_hours
        self.utilization_pct = utilization_pct
        self.tasks = tasks or []
        self.dut_types = dut_types or []
        self.profiles = profiles or []
        self.dut_profile_matrix = dut_profile_matrix or []
        self.team_members = team_members or []
        self.team_size = team_size
        self.has_leader = has_leader
        self.pr_simple = pr_simple
        self.pr_medium = pr_medium
        self.pr_complex = pr_complex
        self.reference_projects = reference_projects or []
        self.risk_flags = risk_flags or []
        self.risk_messages = risk_messages or []
        self.pr_details = pr_details or []


# ── Sheet builders ───────────────────────────────────────

def _build_summary_sheet(ws: Any, data: ExcelReportData) -> None:
    ws.title = "Summary"

    ws.cell(row=1, column=1, value="Test Effort Estimation Report").font = TITLE_FONT
    ws.merge_cells("A1:D1")

    rows = [
        ("Estimation Number", data.estimation_number),
        ("Project Name", data.project_name),
        ("Project Type", data.project_type),
        ("Created By", data.created_by),
        ("Created At", data.created_at),
        ("", ""),
        ("Request Details", ""),
        ("Request Number", data.request_number),
        ("Requester", data.requester_name),
        ("Business Unit", data.business_unit),
        ("Priority", data.priority),
        ("", ""),
        ("Project Parameters", ""),
        ("DUT Count", data.dut_count),
        ("Profile Count", data.profile_count),
        ("DUT × Profile Combinations", data.dut_profile_combinations),
        ("PR Fix Count", data.pr_fix_count),
        ("Expected Delivery", data.expected_delivery),
        ("", ""),
        ("Effort Summary", ""),
        ("Total Tester Hours", f"{data.total_tester_hours:.1f}"),
        ("Test Leader Hours", f"{data.total_leader_hours:.1f}"),
        ("PR Fix Hours", f"{data.pr_fix_hours:.1f}"),
        ("Study Hours", f"{data.study_hours:.1f}"),
        ("Buffer Hours", f"{data.buffer_hours:.1f}"),
        ("Grand Total (Hours)", f"{data.grand_total_hours:.1f}"),
        ("Grand Total (Days)", f"{data.grand_total_days:.1f}"),
        ("", ""),
        ("Feasibility", ""),
        ("Available Capacity (Hours)", f"{data.capacity_hours:.1f}"),
        ("Utilization", f"{data.utilization_pct:.1f}%"),
        ("Status", data.feasibility_status),
    ]

    for i, (label, value) in enumerate(rows, start=3):
        cell_a = ws.cell(row=i, column=1, value=label)
        cell_b = ws.cell(row=i, column=2, value=value)
        if label and value == "":
            cell_a.font = SUBTITLE_FONT
        elif label:
            cell_a.font = LABEL_FONT
            cell_b.font = VALUE_FONT
        if label == "Status":
            cell_b.fill = _feasibility_fill(data.feasibility_status)

    # Risk flags
    risk_row = len(rows) + 4
    if data.risk_messages:
        ws.cell(row=risk_row, column=1, value="Risk Flags").font = SUBTITLE_FONT
        for i, msg in enumerate(data.risk_messages):
            ws.cell(row=risk_row + 1 + i, column=1, value=f"⚠ {msg}").font = VALUE_FONT

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 40


def _build_task_breakdown_sheet(ws: Any, data: ExcelReportData) -> None:
    ws.title = "Task Breakdown"

    headers = ["Task Name", "Type", "Base Hours", "DUT x", "Profile x", "Complexity", "Tester Hours", "Leader Hours", "Study?", "Notes"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    _style_header_row(ws, 1, len(headers))

    for i, task in enumerate(data.tasks, start=2):
        ws.cell(row=i, column=1, value=task.get("task_name", task.get("name", "")))
        ws.cell(row=i, column=2, value=task.get("task_type", ""))
        ws.cell(row=i, column=3, value=task.get("base_hours", 0))
        ws.cell(row=i, column=4, value=task.get("dut_multiplier", 1))
        ws.cell(row=i, column=5, value=task.get("profile_multiplier", 1))
        ws.cell(row=i, column=6, value=task.get("complexity_weight", 1.0))
        ws.cell(row=i, column=7, value=task.get("calculated_hours", 0))
        ws.cell(row=i, column=8, value=task.get("leader_hours", 0))
        ws.cell(row=i, column=9, value="Yes" if task.get("is_new_feature_study") else "")
        ws.cell(row=i, column=10, value=task.get("notes", ""))
        for col in range(1, len(headers) + 1):
            ws.cell(row=i, column=col).border = THIN_BORDER

    # Totals row
    total_row = len(data.tasks) + 2
    ws.cell(row=total_row, column=1, value="TOTAL").font = LABEL_FONT
    ws.cell(row=total_row, column=7, value=data.total_tester_hours).font = LABEL_FONT
    ws.cell(row=total_row, column=8, value=data.total_leader_hours).font = LABEL_FONT

    _auto_width(ws)


def _build_dut_profile_matrix_sheet(ws: Any, data: ExcelReportData) -> None:
    ws.title = "DUT-Profile Matrix"

    if not data.dut_types or not data.profiles:
        ws.cell(row=1, column=1, value="No DUT/Profile data available")
        return

    # Header row: empty corner + profile names
    ws.cell(row=1, column=1, value="DUT \\ Profile").font = LABEL_FONT
    for j, profile in enumerate(data.profiles):
        ws.cell(row=1, column=j + 2, value=profile.get("name", f"Profile {j+1}"))
    _style_header_row(ws, 1, len(data.profiles) + 1)

    # Build set of active combinations
    active_combos = set()
    for combo in data.dut_profile_matrix:
        if len(combo) >= 2:
            active_combos.add((combo[0], combo[1]))

    # Data rows
    for i, dut in enumerate(data.dut_types):
        ws.cell(row=i + 2, column=1, value=dut.get("name", f"DUT {i+1}")).font = LABEL_FONT
        for j, profile in enumerate(data.profiles):
            dut_id = dut.get("id", i + 1)
            prof_id = profile.get("id", j + 1)
            is_active = (dut_id, prof_id) in active_combos or not data.dut_profile_matrix
            cell = ws.cell(row=i + 2, column=j + 2, value="✓" if is_active else "—")
            cell.alignment = Alignment(horizontal="center")
            cell.border = THIN_BORDER
            if is_active:
                cell.fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")

    ws.cell(row=len(data.dut_types) + 3, column=1, value=f"Total combinations: {data.dut_profile_combinations}").font = LABEL_FONT
    _auto_width(ws)


def _build_team_allocation_sheet(ws: Any, data: ExcelReportData) -> None:
    ws.title = "Team Allocation"

    headers = ["Name", "Role", "Hours/Day", "Skills"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    _style_header_row(ws, 1, len(headers))

    if data.team_members:
        for i, member in enumerate(data.team_members, start=2):
            ws.cell(row=i, column=1, value=member.get("name", ""))
            ws.cell(row=i, column=2, value=member.get("role", ""))
            ws.cell(row=i, column=3, value=member.get("available_hours_per_day", 7.0))
            ws.cell(row=i, column=4, value=member.get("skills", ""))
            for col in range(1, len(headers) + 1):
                ws.cell(row=i, column=col).border = THIN_BORDER
    else:
        ws.cell(row=2, column=1, value=f"Team size: {data.team_size} tester(s)")
        ws.cell(row=3, column=1, value=f"Test leader: {'Yes' if data.has_leader else 'No'}")

    summary_row = max(len(data.team_members) + 3, 5)
    ws.cell(row=summary_row, column=1, value="Effort Distribution").font = SUBTITLE_FONT
    ws.cell(row=summary_row + 1, column=1, value="Tester Effort").font = LABEL_FONT
    ws.cell(row=summary_row + 1, column=2, value=f"{data.total_tester_hours:.1f}h")
    ws.cell(row=summary_row + 2, column=1, value="Leader Effort").font = LABEL_FONT
    ws.cell(row=summary_row + 2, column=2, value=f"{data.total_leader_hours:.1f}h")

    _auto_width(ws)


def _build_pr_fixes_sheet(ws: Any, data: ExcelReportData) -> None:
    ws.title = "PR Fixes"

    headers = ["Complexity", "Count", "Hours Each", "Subtotal"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    _style_header_row(ws, 1, len(headers))

    rows = [
        ("Simple", data.pr_simple, 2, data.pr_simple * 2),
        ("Medium", data.pr_medium, 4, data.pr_medium * 4),
        ("Complex", data.pr_complex, 8, data.pr_complex * 8),
    ]
    for i, (complexity, count, hours, subtotal) in enumerate(rows, start=2):
        ws.cell(row=i, column=1, value=complexity)
        ws.cell(row=i, column=2, value=count)
        ws.cell(row=i, column=3, value=hours)
        ws.cell(row=i, column=4, value=subtotal)
        for col in range(1, 5):
            ws.cell(row=i, column=col).border = THIN_BORDER

    ws.cell(row=5, column=1, value="TOTAL (before DUT scaling)").font = LABEL_FONT
    ws.cell(row=5, column=2, value=data.pr_fix_count).font = LABEL_FONT
    base_total = data.pr_simple * 2 + data.pr_medium * 4 + data.pr_complex * 8
    ws.cell(row=5, column=4, value=base_total).font = LABEL_FONT

    ws.cell(row=7, column=1, value=f"DUT count: {data.dut_count}").font = LABEL_FONT
    ws.cell(row=8, column=1, value=f"Total PR fix effort (× {data.dut_count} DUTs):").font = LABEL_FONT
    ws.cell(row=8, column=2, value=f"{data.pr_fix_hours:.1f}h").font = LABEL_FONT

    # PR details table (optional)
    if data.pr_details:
        detail_start = 10
        ws.cell(row=detail_start, column=1, value="PR Details").font = SUBTITLE_FONT
        detail_headers = ["PR Number", "Link", "Complexity", "Status"]
        for col, header in enumerate(detail_headers, 1):
            ws.cell(row=detail_start + 1, column=col, value=header)
        _style_header_row(ws, detail_start + 1, len(detail_headers))

        for i, pr in enumerate(data.pr_details, start=detail_start + 2):
            ws.cell(row=i, column=1, value=pr.get("pr_number", ""))
            ws.cell(row=i, column=2, value=pr.get("link", ""))
            ws.cell(row=i, column=3, value=pr.get("complexity", ""))
            ws.cell(row=i, column=4, value=pr.get("status", ""))
            for col in range(1, 5):
                ws.cell(row=i, column=col).border = THIN_BORDER

    _auto_width(ws)


def _build_reference_data_sheet(ws: Any, data: ExcelReportData) -> None:
    ws.title = "Reference Data"

    if not data.reference_projects:
        ws.cell(row=1, column=1, value="No reference projects linked")
        return

    headers = ["Project Name", "Type", "Estimated Hours", "Actual Hours", "Accuracy Ratio", "DUTs", "Profiles", "PRs"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    _style_header_row(ws, 1, len(headers))

    for i, proj in enumerate(data.reference_projects, start=2):
        est = proj.get("estimated_hours", 0) or 0
        act = proj.get("actual_hours", 0) or 0
        ratio = act / est if est > 0 else 0

        ws.cell(row=i, column=1, value=proj.get("project_name", ""))
        ws.cell(row=i, column=2, value=proj.get("project_type", ""))
        ws.cell(row=i, column=3, value=est)
        ws.cell(row=i, column=4, value=act)
        ws.cell(row=i, column=5, value=f"{ratio:.2f}")
        ws.cell(row=i, column=6, value=proj.get("dut_count", ""))
        ws.cell(row=i, column=7, value=proj.get("profile_count", ""))
        ws.cell(row=i, column=8, value=proj.get("pr_count", ""))
        for col in range(1, len(headers) + 1):
            ws.cell(row=i, column=col).border = THIN_BORDER

    _auto_width(ws)


# ── Main generation function ─────────────────────────────

def generate_excel_report(data: ExcelReportData, output_path: str | Path | None = None) -> bytes | None:
    """Generate an Excel report and either save to file or return bytes.

    Args:
        data: ExcelReportData with all estimation information.
        output_path: If provided, saves to file and returns None.
                     If None, returns the workbook as bytes.
    """
    wb = Workbook()

    # Sheet 1 - Summary (uses the default sheet)
    _build_summary_sheet(wb.active, data)

    # Sheet 2 - Task Breakdown
    _build_task_breakdown_sheet(wb.create_sheet(), data)

    # Sheet 3 - DUT-Profile Matrix
    _build_dut_profile_matrix_sheet(wb.create_sheet(), data)

    # Sheet 4 - Team Allocation
    _build_team_allocation_sheet(wb.create_sheet(), data)

    # Sheet 5 - PR Fixes
    _build_pr_fixes_sheet(wb.create_sheet(), data)

    # Sheet 6 - Reference Data
    _build_reference_data_sheet(wb.create_sheet(), data)

    if output_path:
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(str(output_path))
        return None
    else:
        buffer = BytesIO()
        wb.save(buffer)
        return buffer.getvalue()
