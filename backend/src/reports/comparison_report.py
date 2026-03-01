"""Side-by-side comparison report for two estimations."""

import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .templates import ComparisonReportData


def generate_comparison_excel(data: ComparisonReportData) -> bytes:
    """Generate Excel report comparing two estimations side by side."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Comparison"

    # Styles
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    label_font = Font(bold=True, size=10)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    est_a = data.estimation_a
    est_b = data.estimation_b

    # Title
    ws.merge_cells("A1:D1")
    title_cell = ws["A1"]
    title_cell.value = "Estimation Comparison Report"
    title_cell.font = Font(bold=True, size=14, color="2F5496")
    title_cell.alignment = Alignment(horizontal="center")

    # Confidentiality notice
    ws.merge_cells("A2:D2")
    ws["A2"].value = data.metadata.confidentiality_notice
    ws["A2"].font = Font(italic=True, size=8, color="999999")
    ws["A2"].alignment = Alignment(horizontal="center")

    # Headers
    row = 4
    headers = ["Metric", est_a.get("estimation_number", "Estimation A"),
               est_b.get("estimation_number", "Estimation B"), "Difference"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")

    # Comparison metrics
    metrics = [
        ("Project Name", "project_name", "project_name", False),
        ("Project Type", "project_type", "project_type", False),
        ("Grand Total (hours)", "grand_total_hours", "grand_total_hours", True),
        ("Grand Total (days)", "grand_total_days", "grand_total_days", True),
        ("Tester Hours", "total_tester_hours", "total_tester_hours", True),
        ("Leader Hours", "total_leader_hours", "total_leader_hours", True),
        ("Feasibility", "feasibility_status", "feasibility_status", False),
        ("DUT Count", "dut_count", "dut_count", True),
        ("Profile Count", "profile_count", "profile_count", True),
        ("DUT×Profile Combinations", "dut_profile_combinations", "dut_profile_combinations", True),
        ("PR Fix Count", "pr_fix_count", "pr_fix_count", True),
        ("Status", "status", "status", False),
    ]

    for label, key_a, key_b, is_numeric in metrics:
        row += 1
        val_a = est_a.get(key_a, "")
        val_b = est_b.get(key_b, "")

        ws.cell(row=row, column=1, value=label).font = label_font
        ws.cell(row=row, column=2, value=val_a)
        ws.cell(row=row, column=3, value=val_b)

        if is_numeric and isinstance(val_a, (int, float)) and isinstance(val_b, (int, float)):
            diff = val_b - val_a
            pct = f" ({diff / val_a * 100:+.1f}%)" if val_a != 0 else ""
            ws.cell(row=row, column=4, value=f"{diff:+.1f}{pct}")
        else:
            ws.cell(row=row, column=4, value="—" if val_a == val_b else "Changed")

        for col in range(1, 5):
            ws.cell(row=row, column=col).border = thin_border

    # Task comparison section
    row += 2
    ws.merge_cells(f"A{row}:D{row}")
    ws.cell(row=row, column=1, value="Task Breakdown Comparison").font = Font(bold=True, size=12, color="2F5496")

    row += 1
    task_headers = ["Task Name", "Hours (A)", "Hours (B)", "Difference"]
    for col, header in enumerate(task_headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border

    tasks_a = {t.get("task_name", t.get("name", "")): t for t in est_a.get("tasks", [])}
    tasks_b = {t.get("task_name", t.get("name", "")): t for t in est_b.get("tasks", [])}
    all_task_names = list(dict.fromkeys(list(tasks_a.keys()) + list(tasks_b.keys())))

    for task_name in all_task_names:
        row += 1
        hours_a = tasks_a.get(task_name, {}).get("calculated_hours", 0)
        hours_b = tasks_b.get(task_name, {}).get("calculated_hours", 0)
        diff = hours_b - hours_a

        ws.cell(row=row, column=1, value=task_name)
        ws.cell(row=row, column=2, value=round(hours_a, 1))
        ws.cell(row=row, column=3, value=round(hours_b, 1))
        ws.cell(row=row, column=4, value=f"{diff:+.1f}")

        for col in range(1, 5):
            ws.cell(row=row, column=col).border = thin_border

    # Auto-width columns
    for col in range(1, 5):
        ws.column_dimensions[get_column_letter(col)].width = 25

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
