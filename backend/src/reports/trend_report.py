"""Historical trend analysis report with charts."""

import io
from typing import Optional

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .templates import TrendReportData


def _create_accuracy_chart(buf: io.BytesIO, projects: list[dict]) -> Optional[bytes]:
    """Create accuracy trend chart using matplotlib."""
    try:
        import matplotlib
        matplotlib.use("Agg")
        import matplotlib.pyplot as plt

        if not projects:
            return None

        names = [p.get("project_name", f"P{i}") for i, p in enumerate(projects)]
        estimated = [p.get("estimated_hours", 0) or 0 for p in projects]
        actual = [p.get("actual_hours", 0) or 0 for p in projects]

        fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(14, 5))

        # Bar chart: estimated vs actual
        x = range(len(names))
        width = 0.35
        ax1.bar([i - width/2 for i in x], estimated, width, label="Estimated", color="#2F5496")
        ax1.bar([i + width/2 for i in x], actual, width, label="Actual", color="#E07020")
        ax1.set_xlabel("Project")
        ax1.set_ylabel("Hours")
        ax1.set_title("Estimated vs Actual Hours")
        ax1.set_xticks(list(x))
        ax1.set_xticklabels(names, rotation=45, ha="right", fontsize=8)
        ax1.legend()

        # Line chart: accuracy ratio over time
        ratios = [a / e if e > 0 else 1.0 for e, a in zip(estimated, actual)]
        ax2.plot(names, ratios, marker="o", color="#2F5496", linewidth=2)
        ax2.axhline(y=1.0, color="#198754", linestyle="--", label="Perfect accuracy")
        ax2.axhline(y=1.3, color="#dc3545", linestyle="--", label="Underestimate threshold")
        ax2.set_xlabel("Project")
        ax2.set_ylabel("Accuracy Ratio (Actual/Estimated)")
        ax2.set_title("Estimation Accuracy Trend")
        ax2.set_xticklabels(names, rotation=45, ha="right", fontsize=8)
        ax2.legend()

        plt.tight_layout()
        chart_buf = io.BytesIO()
        plt.savefig(chart_buf, format="png", dpi=150, bbox_inches="tight")
        plt.close(fig)
        return chart_buf.getvalue()
    except ImportError:
        return None


def generate_trend_excel(data: TrendReportData) -> bytes:
    """Generate trend analysis Excel report with embedded charts."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Trend Analysis"

    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    # Title
    ws.merge_cells("A1:G1")
    ws["A1"].value = "Historical Trend Analysis Report"
    ws["A1"].font = Font(bold=True, size=14, color="2F5496")
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:G2")
    ws["A2"].value = data.metadata.confidentiality_notice
    ws["A2"].font = Font(italic=True, size=8, color="999999")
    ws["A2"].alignment = Alignment(horizontal="center")

    # Data table
    row = 4
    headers = ["Project", "Type", "Estimated (h)", "Actual (h)", "Accuracy Ratio",
               "DUT Count", "Completion Date"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")

    for proj in data.projects:
        row += 1
        est_h = proj.get("estimated_hours", 0) or 0
        act_h = proj.get("actual_hours", 0) or 0
        ratio = round(act_h / est_h, 2) if est_h > 0 else 0

        values = [
            proj.get("project_name", ""),
            proj.get("project_type", ""),
            round(est_h, 1),
            round(act_h, 1),
            ratio,
            proj.get("dut_count", 0),
            proj.get("completion_date", ""),
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = thin_border

    # Summary statistics
    data_end_row = row
    row += 2
    ws.cell(row=row, column=1, value="Summary Statistics").font = Font(bold=True, size=12, color="2F5496")

    projects = data.projects
    if projects:
        est_hours = [p.get("estimated_hours", 0) or 0 for p in projects]
        act_hours = [p.get("actual_hours", 0) or 0 for p in projects]
        ratios = [a / e if e > 0 else 1.0 for e, a in zip(est_hours, act_hours)]

        stats = [
            ("Total Projects", len(projects)),
            ("Avg Estimated Hours", round(sum(est_hours) / len(est_hours), 1) if est_hours else 0),
            ("Avg Actual Hours", round(sum(act_hours) / len(act_hours), 1) if act_hours else 0),
            ("Avg Accuracy Ratio", round(sum(ratios) / len(ratios), 2) if ratios else 0),
            ("Best Accuracy", round(min(ratios), 2) if ratios else 0),
            ("Worst Accuracy", round(max(ratios), 2) if ratios else 0),
        ]
        for label, value in stats:
            row += 1
            ws.cell(row=row, column=1, value=label).font = Font(bold=True)
            ws.cell(row=row, column=2, value=value)

    # Add Excel chart
    if len(data.projects) >= 2:
        chart = BarChart()
        chart.type = "col"
        chart.title = "Estimated vs Actual Hours"
        chart.y_axis.title = "Hours"
        chart.x_axis.title = "Project"
        chart.width = 20
        chart.height = 12

        data_ref_est = Reference(ws, min_col=3, min_row=4, max_row=data_end_row)
        data_ref_act = Reference(ws, min_col=4, min_row=4, max_row=data_end_row)
        cats = Reference(ws, min_col=1, min_row=5, max_row=data_end_row)
        chart.add_data(data_ref_est, titles_from_data=True)
        chart.add_data(data_ref_act, titles_from_data=True)
        chart.set_categories(cats)
        ws.add_chart(chart, f"A{row + 3}")

    # Auto-width
    for col in range(1, 8):
        ws.column_dimensions[get_column_letter(col)].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def generate_trend_chart_png(data: TrendReportData) -> bytes | None:
    """Generate standalone trend chart as PNG image."""
    return _create_accuracy_chart(io.BytesIO(), data.projects)
